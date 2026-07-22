from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aso_logic import (
    AsoResult,
    ChemistrySettings,
    DisplaySpan,
    IdtConversionRow,
    PenaltyAsoResult,
    chemistry_display_label,
    header_display_positions_5to3,
    mutation_header_indexes,
    normalise_mutation_type,
)


BLUE = "1F4E79"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "C6EFCE"
LIGHT_RED = "FCE4D6"
LIGHT_YELLOW = "FFF2CC"
GRID_GREY = "F3F3F3"
MID_GREY = "BFBFBF"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN_TEXT = "006100"
RED = "C00000"


def _thin_border(color: str = "D9D9D9") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _rich_display_text(text: str, spans: tuple[DisplaySpan, ...]) -> CellRichText | str:
    if not spans:
        return text

    def run(value: str, color: str) -> TextBlock:
        return TextBlock(InlineFont(rFont="Arial", sz=12, color=color), value)

    rich = CellRichText()
    pos = 0
    for span in sorted(spans, key=lambda item: item.start):
        if span.start > pos:
            rich.append(run(text[pos : span.start], BLACK))
        color = RED if span.kind == "mutation" else BLACK
        rich.append(run(text[span.start : span.end], color))
        pos = span.end
    if pos < len(text):
        rich.append(run(text[pos:], BLACK))
    return rich


def _alignment_boundaries(result: AsoResult) -> list[int]:
    mutation_type = normalise_mutation_type(result.inputs.mutation_type)
    if mutation_type == "DELETION":
        boundary = result.mutation_start_reversed - result.crop_start
        if 0 <= boundary <= len(result.header_positions):
            return [boundary]
        return []

    if mutation_type in {"INSERTION", "SUBSTITUTION"}:
        highlighted = mutation_header_indexes(result)
        if highlighted:
            return [min(highlighted), max(highlighted) + 1]
    return []


def _apply_vertical_marker(ws, row_start: int, row_end: int, grid_start_col: int, grid_width: int, boundary: int) -> None:
    marker_side = Side(style="medium", color=BLACK)
    thin_side = Side(style="thin", color="BFBFBF")

    if boundary >= grid_width:
        col = grid_start_col + grid_width - 1
        side_name = "right"
    else:
        col = grid_start_col + max(0, boundary)
        side_name = "left"

    for row in range(row_start, row_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(
            left=marker_side if side_name == "left" else cell.border.left or thin_side,
            right=marker_side if side_name == "right" else cell.border.right or thin_side,
            top=cell.border.top or thin_side,
            bottom=cell.border.bottom or thin_side,
        )


def export_result_to_xlsx(result: AsoResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "ASO Output"
    ws.sheet_view.showGridLines = True

    headers = [
        "ASO ID",
        "Sequence 5' to 3' with IDT Codes",
        "Highlighted Sequence 5' to 3'",
        "Variant Start Position",
        "Chemistry",
    ]
    table_header_row = 2
    position_row = 1
    rna_row = 2
    first_data_row = 3
    grid_start_col = len(headers) + 2
    display_positions = header_display_positions_5to3(result)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_header_row, column=col, value=header)
        cell.font = Font(name="Arial", size=12, color=BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border()

    end_label_cell = ws.cell(row=table_header_row, column=grid_start_col - 1, value="3'")
    end_label_cell.font = Font(name="Arial", size=12, bold=True, color=BLACK)
    end_label_cell.alignment = Alignment(horizontal="center", vertical="center")

    highlighted_header_indexes = mutation_header_indexes(result)
    for idx, pos in enumerate(display_positions, start=grid_start_col):
        top = ws.cell(row=position_row, column=idx, value=pos)
        base = ws.cell(row=rna_row, column=idx, value=result.header_bases[idx - grid_start_col])
        for cell in (top, base):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border()
        top.font = Font(name="Aptos Narrow", size=12, color=BLACK)
        if idx - grid_start_col in highlighted_header_indexes:
            base.fill = PatternFill("solid", fgColor=LIGHT_RED)
            base.font = Font(name="Aptos Narrow", size=12, color=RED)
        else:
            base.font = Font(name="Aptos Narrow", size=12, color=BLACK)

    for row_offset, aso_row in enumerate(result.rows, start=first_data_row):
        values = [
            aso_row.aso_id,
            aso_row.idt_code,
            None,
            aso_row.starting_position,
            aso_row.chemistry,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col, value=value)
            cell.font = Font(name="Arial", size=12, color=BLACK)
            cell.border = Border()
            cell.alignment = Alignment(vertical="top", wrap_text=(col in {2, 3, 5}))
        display_cell = ws.cell(row=row_offset, column=3)
        display_cell.value = _rich_display_text(aso_row.display_sequence, aso_row.display_spans)
        display_cell.border = Border()
        display_cell.alignment = Alignment(vertical="top")

        for idx, cell_value in enumerate(aso_row.grid_cells, start=grid_start_col):
            used = cell_value != "##"
            cell = ws.cell(row=row_offset, column=idx, value=cell_value if used else "")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border("BFBFBF")
            if used:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
                cell.font = Font(name="Aptos Narrow", size=12, color=GREEN_TEXT)
            else:
                cell.fill = PatternFill("solid", fgColor=GRID_GREY)
                cell.font = Font(name="Aptos Narrow", size=12, color=MID_GREY)

    for col, width in {
        "A": 24,
        "B": 58,
        "C": 30,
        "D": 28,
        "E": 42,
        "F": 5,
    }.items():
        ws.column_dimensions[col].width = width
    grid_end = grid_start_col + len(result.header_positions) - 1
    grid_col_width = max(5, len(str(max(display_positions, default=1))) + 2)
    for idx in range(grid_start_col, grid_end + 1):
        ws.column_dimensions[get_column_letter(idx)].width = grid_col_width

    for row in range(1, first_data_row + len(result.rows)):
        ws.row_dimensions[row].height = 24

    last_data_row = first_data_row + len(result.rows) - 1
    for boundary in _alignment_boundaries(result):
        _apply_vertical_marker(ws, position_row, last_data_row, grid_start_col, len(result.header_positions), boundary)

    ws.freeze_panes = None

    wb.save(output_path)
    return output_path


def _content_width(values: list[object], *, minimum: int = 10, maximum: int = 120) -> int:
    longest = max((len(str(value)) for value in values if value is not None), default=0)
    return max(minimum, min(maximum, longest + 3))


def export_idt_conversion_to_xlsx(
    rows: tuple[IdtConversionRow, ...] | list[IdtConversionRow],
    chemistry: ChemistrySettings,
    output_path: str | Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "IDT Converter"
    ws.sheet_view.showGridLines = True

    title = ws.cell(row=1, column=1, value="IDT notation converter")
    title.font = Font(name="Arial", size=14, bold=True, color=BLUE)

    chemistry_row = ws.cell(row=2, column=1, value="Chemistry")
    chemistry_row.font = Font(name="Arial", size=12, bold=True, color=BLACK)
    ws.cell(row=2, column=2, value=chemistry_display_label(chemistry))

    length_row = ws.cell(row=3, column=1, value="Expected length")
    length_row.font = Font(name="Arial", size=12, bold=True, color=BLACK)
    ws.cell(row=3, column=2, value=f"{chemistry.aso_length} bases")

    headers = ["No.", "Sequence 5' to 3'", "Clean sequence", "IDT notation"]
    header_row = 5
    first_data_row = header_row + 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(name="Arial", size=12, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border(BLUE)

    for row_offset, conversion_row in enumerate(rows, start=first_data_row):
        values = [
            conversion_row.row_number,
            conversion_row.input_sequence,
            conversion_row.clean_sequence,
            conversion_row.idt_code,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col, value=value)
            cell.font = Font(name="Arial", size=12, color=BLACK)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 4))

    row_values = list(rows)
    chemistry_label = chemistry_display_label(chemistry)
    widths = [
        _content_width(["No.", *[row.row_number for row in row_values]], minimum=7, maximum=12),
        _content_width(
            ["Sequence 5' to 3'", chemistry_label, *[row.input_sequence for row in row_values]],
            minimum=18,
            maximum=100,
        ),
        _content_width(["Clean sequence", *[row.clean_sequence for row in row_values]], minimum=18, maximum=80),
        _content_width(["IDT notation", *[row.idt_code for row in row_values]], minimum=32, maximum=220),
    ]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[header_row].height = 24
    for row_idx in range(first_data_row, first_data_row + len(row_values)):
        ws.row_dimensions[row_idx].height = 30

    ws.freeze_panes = None

    wb.save(output_path)
    return output_path


def export_penalty_design_to_xlsx(result: PenaltyAsoResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Penalty ASOs"
    ws.sheet_view.showGridLines = True

    title = ws.cell(row=1, column=1, value="Penalty ASO design")
    title.font = Font(name="Arial", size=14, bold=True, color=BLUE)

    chemistry_cell = ws.cell(row=2, column=1, value="Chemistry")
    chemistry_cell.font = Font(name="Arial", size=12, bold=True, color=BLACK)
    ws.cell(row=2, column=2, value=chemistry_display_label(result.chemistry))

    settings_cell = ws.cell(row=3, column=1, value="Penalty settings")
    settings_cell.font = Font(name="Arial", size=12, bold=True, color=BLACK)
    ws.cell(
        row=3,
        column=2,
        value=f"{result.inputs.penalty_position_mode}; {result.inputs.penalty_base_mode}",
    )

    headers = [
        "No.",
        "Parent ASO ID",
        "Penalty ASO ID",
        "Parent Start Position",
        "Penalty ASO Position",
        "RNA 3' to 5' Position",
        "RNA 5' to 3' Position",
        "RNA Context",
        "Target Base",
        "Canonical ASO Base",
        "Penalty ASO Base",
        "Mismatch Pair",
        "Priority",
        "Score",
        "Reason",
        "Sequence 5' to 3'",
        "IDT notation",
        "Chemistry",
    ]
    header_row = 5
    first_data_row = header_row + 1
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(name="Arial", size=12, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border(BLUE)

    priority_fills = {
        "Recommended": PatternFill("solid", fgColor=LIGHT_GREEN),
        "Alternative": PatternFill("solid", fgColor=LIGHT_YELLOW),
        "Lower priority": PatternFill("solid", fgColor=LIGHT_RED),
    }

    for row_offset, penalty_row in enumerate(result.rows, start=first_data_row):
        values = [
            penalty_row.row_number,
            penalty_row.parent_aso_id,
            penalty_row.penalty_aso_id,
            penalty_row.parent_start,
            penalty_row.penalty_aso_position,
            penalty_row.target_position_3to5,
            penalty_row.target_position_5to3,
            penalty_row.local_rna_context,
            penalty_row.target_base,
            penalty_row.canonical_aso_base,
            penalty_row.penalty_aso_base,
            penalty_row.mismatch_pair,
            penalty_row.priority,
            penalty_row.score,
            penalty_row.reason,
            penalty_row.clean_sequence,
            penalty_row.idt_code,
            penalty_row.chemistry,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_offset, column=col, value=value)
            cell.font = Font(name="Arial", size=12, color=BLACK)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=(col in {15, 17, 18}))
            if col == 13:
                cell.fill = priority_fills.get(str(value), PatternFill())

    row_values = list(result.rows)
    for idx, header in enumerate(headers, start=1):
        values = [getattr(row, _penalty_export_attr(idx), "") for row in row_values]
        ws.column_dimensions[get_column_letter(idx)].width = _content_width(
            [header, *values],
            minimum=8,
            maximum=220 if idx == 17 else 90,
        )

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[header_row].height = 28
    for row_idx in range(first_data_row, first_data_row + len(row_values)):
        ws.row_dimensions[row_idx].height = 32

    ws.freeze_panes = None

    wb.save(output_path)
    return output_path


def _penalty_export_attr(column_index: int) -> str:
    return [
        "row_number",
        "parent_aso_id",
        "penalty_aso_id",
        "parent_start",
        "penalty_aso_position",
        "target_position_3to5",
        "target_position_5to3",
        "local_rna_context",
        "target_base",
        "canonical_aso_base",
        "penalty_aso_base",
        "mismatch_pair",
        "priority",
        "score",
        "reason",
        "clean_sequence",
        "idt_code",
        "chemistry",
    ][column_index - 1]
