import unittest
from tempfile import TemporaryDirectory
from pathlib import Path

from openpyxl import load_workbook

from aso_logic import (
    AsoInputs,
    PenaltyAsoInputs,
    chemistry_optimization_walk,
    combine_base_modification,
    convert_sequences_to_idt,
    complement_base,
    generate_design,
    generate_penalty_design,
    header_display_positions_5to3,
    idt_aso_custom,
    idt_aso_per_position,
    mutation_header_indexes,
    resolve_chemistry,
)
from excel_export import export_idt_conversion_to_xlsx, export_penalty_design_to_xlsx


class AsoLogicTests(unittest.TestCase):
    def test_lna_idt_code_matches_vba_token_rules(self):
        code = idt_aso_custom("AUGCAUGCAUGCAUGCAU", 12, 3, "LNA", "PS")
        self.assertTrue(code.startswith("+A*+T*+G*C*"))
        self.assertTrue(code.endswith("*+C*+A*+T"))
        self.assertEqual(code.count("*"), 17)

    def test_moe_idt_code_uses_terminal_and_internal_moe_tokens(self):
        code = idt_aso_custom("AUGCAUGCAUGCAUGCAU", 10, 4, "MOE", "PO")
        self.assertTrue(code.startswith("/52MOErA//i2MOErT//i2MOErG//i2MOErC/"))
        self.assertTrue(code.endswith("/i2MOErG//i2MOErC//i2MOErA//32MOErT/"))
        self.assertNotIn("*", code)

    def test_per_position_idt_code_supports_ome_fluoro_and_5mec(self):
        code = idt_aso_per_position(
            "ACCU",
            ("2'OMe", "2'F", "DNA + 5MeC", "MOE"),
            ("PS", "PO", "PS"),
        )
        self.assertEqual(code, "mA*/i2FC//iMe-dC/*/32MOErT/")

    def test_per_position_chemistry_can_combine_ribose_and_5mec(self):
        self.assertEqual(combine_base_modification("LNA", "5MeC"), "LNA + 5MeC")
        code = idt_aso_per_position(
            "ACCU",
            ("LNA + 5MeC", "MOE + 5MeC", "2'OMe + 5MeC", "2'F + 5MeC"),
            ("PS", "PO", "PS"),
        )
        self.assertEqual(code, "+A*/i2MOErC/mC*/32FU/")

    def test_unmodified_dna_preset_uses_plain_dna_po_chemistry(self):
        chemistry = resolve_chemistry(AsoInputs(aso_chemistry="Unmodified DNA"))

        self.assertEqual(chemistry.aso_length, 20)
        self.assertEqual(set(chemistry.base_modifications), {"DNA"})
        self.assertEqual(set(chemistry.linkages), {"PO"})
        self.assertEqual(
            idt_aso_per_position("AUGCAUGCAUGCAUGCAUGC", chemistry.base_modifications, chemistry.linkages),
            "ATGCATGCATGCATGCATGC",
        )

    def test_kt777_preset_uses_mixed_linkages_and_5mec_gap(self):
        chemistry = resolve_chemistry(AsoInputs(aso_chemistry="KT777/valeriasen"))
        self.assertEqual(chemistry.aso_length, 20)
        self.assertEqual(chemistry.linkages.count("PO"), 6)
        self.assertEqual(chemistry.linkages[0], "PS")
        self.assertEqual(chemistry.linkages[-1], "PS")

        code = idt_aso_per_position("AAAAACCCCCAAAAACCCCC", chemistry.base_modifications, chemistry.linkages)
        self.assertEqual(code.count("*"), 13)
        self.assertIn("/iMe-dC/", code)

    def test_convert_sequences_to_idt_handles_multiline_input(self):
        chemistry = resolve_chemistry(AsoInputs(aso_chemistry="3-10-3 LNA/DNA"))
        rows = convert_sequences_to_idt(
            "AUGCAUGCAUGCAUGC\nAUGC-AUGCAUGCAUGC\n",
            chemistry,
        )
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].clean_sequence, "AUGCAUGCAUGCAUGC")
        self.assertTrue(rows[0].idt_code.startswith("+A*+T*+G*"))
        self.assertEqual(rows[1].clean_sequence, "AUGCAUGCAUGCAUGC")

    def test_export_idt_conversion_to_xlsx_writes_dynamic_widths(self):
        chemistry = resolve_chemistry(AsoInputs(aso_chemistry="KT777/valeriasen"))
        rows = convert_sequences_to_idt(
            "AUGCAUGCAUGCAUGCAUGC\nAUGC-AUGCAUGCAUGCAUGC\n",
            chemistry,
        )
        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "idt_converter.xlsx"
            export_idt_conversion_to_xlsx(rows, chemistry, output_path)
            wb = load_workbook(output_path)
            ws = wb["IDT Converter"]

            self.assertEqual(ws["A1"].value, "IDT notation converter")
            self.assertEqual(ws["A5"].value, "No.")
            self.assertEqual(ws["D5"].value, "IDT notation")
            self.assertEqual(ws["C6"].value, "AUGCAUGCAUGCAUGCAUGC")
            self.assertIn("/iMe-dC/", ws["D6"].value)
            self.assertGreater(ws.column_dimensions["D"].width, ws.column_dimensions["C"].width)

    def test_chemistry_optimization_walk_slides_motif_only_inside_core(self):
        chemistry = resolve_chemistry(AsoInputs(aso_chemistry="5-10-5 MOE/DNA"))
        base_mods = list(chemistry.base_modifications)
        base_mods[5] = "MOE"
        base_mods[6] = "MOE"

        rows = chemistry_optimization_walk(
            "AUGCAUGCAUGCAUGCAUGC",
            tuple(base_mods),
            chemistry.linkages,
            chemistry.wing_length,
            chemistry.gap_length,
            (5, 6),
        )

        self.assertEqual(len(rows), 9)
        self.assertEqual(rows[0].motif_aso_positions, (6, 7))
        self.assertEqual(rows[0].motif_gap_positions, (1, 2))
        self.assertEqual(rows[1].motif_aso_positions, (7, 8))
        self.assertEqual(rows[-1].motif_aso_positions, (14, 15))
        self.assertEqual(rows[1].base_modifications[5], "DNA")
        self.assertEqual(rows[1].base_modifications[6], "MOE")
        self.assertEqual(rows[1].base_modifications[7], "MOE")
        self.assertEqual(rows[1].base_modifications[:5], chemistry.base_modifications[:5])
        self.assertEqual(rows[1].base_modifications[15:], chemistry.base_modifications[15:])

    def test_chemistry_optimization_walk_can_skip_by_step_size(self):
        chemistry = resolve_chemistry(AsoInputs(aso_chemistry="5-10-5 MOE/DNA"))
        base_mods = list(chemistry.base_modifications)
        base_mods[5] = "MOE"
        base_mods[6] = "MOE"

        rows = chemistry_optimization_walk(
            "AUGCAUGCAUGCAUGCAUGC",
            tuple(base_mods),
            chemistry.linkages,
            chemistry.wing_length,
            chemistry.gap_length,
            (5, 6),
            step_size=2,
        )

        self.assertEqual(len(rows), 5)
        self.assertEqual([row.motif_gap_positions for row in rows], [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10)])

    def test_penalty_design_defaults_to_like_for_like_candidate(self):
        result = generate_penalty_design(
            PenaltyAsoInputs(
                penalty_position_mode="Selected ASO positions",
                selected_penalty_positions="4",
                parent_count=1,
                rna_sequence="AAAAAAAAAAAAAAAAAAAAUAAA",
            )
        )

        self.assertEqual(len(result.rows), 1)
        row = result.rows[0]
        self.assertEqual(row.target_base, "U")
        self.assertEqual(row.canonical_aso_base, "A")
        self.assertEqual(row.penalty_aso_base, "U")
        self.assertEqual(row.mismatch_pair, "U:U")

    def test_penalty_design_ranks_wobble_like_candidates_lower(self):
        result = generate_penalty_design(
            PenaltyAsoInputs(
                target_gene="GENE",
                target_identifier="Penalty",
                chemistry_number="C1",
                penalty_position_mode="Selected ASO positions",
                selected_penalty_positions="4",
                penalty_base_mode="All mismatches",
                parent_count=1,
                rna_sequence="AAAAAAAAAAAAAAAAAAAAUAAA",
            )
        )

        self.assertEqual(len(result.rows), 3)
        self.assertEqual({row.penalty_aso_base for row in result.rows}, {"C", "G", "U"})
        wobble_row = next(row for row in result.rows if row.mismatch_pair == "G:U")
        self.assertEqual(wobble_row.priority, "Lower priority")
        self.assertIn("wobble", wobble_row.reason)
        self.assertTrue(wobble_row.idt_code)
        self.assertFalse(wobble_row.idt_code.startswith("#ERROR"))

    def test_penalty_design_non_wobble_mode_filters_wobble_candidates(self):
        result = generate_penalty_design(
            PenaltyAsoInputs(
                penalty_position_mode="Selected ASO positions",
                selected_penalty_positions="4",
                penalty_base_mode="Non-wobble only",
                parent_count=1,
                rna_sequence="AAAAAAAAAAAAAAAAAAAAUAAA",
            )
        )

        self.assertEqual({row.mismatch_pair for row in result.rows}, {"C:U", "U:U"})

    def test_export_penalty_design_to_xlsx_writes_ranked_rows(self):
        result = generate_penalty_design(
            PenaltyAsoInputs(
                target_gene="GENE",
                target_identifier="Penalty",
                chemistry_number="C1",
                penalty_position_mode="Selected ASO positions",
                selected_penalty_positions="4",
                penalty_base_mode="All mismatches",
                parent_count=1,
                rna_sequence="AAAAAAAAAAAAAAAAAAAAUAAA",
            )
        )

        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "penalty.xlsx"
            export_penalty_design_to_xlsx(result, output_path)
            wb = load_workbook(output_path)
            ws = wb["Penalty ASOs"]

            self.assertEqual(ws["A1"].value, "Penalty ASO design")
            self.assertEqual(ws["C5"].value, "Penalty ASO ID")
            self.assertEqual(ws["L6"].value, result.rows[0].mismatch_pair)
            self.assertEqual(ws["Q5"].value, "IDT notation")
            self.assertTrue(ws["Q6"].value)

    def test_chemistry_optimization_walk_rejects_wing_motif(self):
        chemistry = resolve_chemistry(AsoInputs(aso_chemistry="5-10-5 MOE/DNA"))
        with self.assertRaisesRegex(ValueError, "central core"):
            chemistry_optimization_walk(
                "AUGCAUGCAUGCAUGCAUGC",
                chemistry.base_modifications,
                chemistry.linkages,
                chemistry.wing_length,
                chemistry.gap_length,
                (4, 5),
            )

    def test_fixed_chemistry_display_includes_backbone(self):
        sequence = "AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA"
        result = generate_design(
            AsoInputs(
                mutation_type="Insertion",
                mutation_length=1,
                mutation_start=22,
                rna_sequence=sequence,
            )
        )
        self.assertEqual(result.rows[0].chemistry, "3-12-3 LNA/DNA, PS backbone modification")

    def test_custom_chemistry_display_includes_dimensions_and_backbone(self):
        sequence = "AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA"
        result = generate_design(
            AsoInputs(
                aso_chemistry="Custom",
                gap_length=10,
                wing_length=4,
                wing_chemistry="MOE",
                backbone_modification="PO",
                mutation_type="Insertion",
                mutation_length=1,
                mutation_start=22,
                rna_sequence=sequence,
            )
        )
        self.assertEqual(result.rows[0].chemistry, "Custom, 4-10-4 MOE/DNA, PO backbone modification")

    def test_insertion_microwalk_matches_workbook_window_math(self):
        sequence = "AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA"
        result = generate_design(
            AsoInputs(
                target_gene="GENE",
                snp_identifier="Example",
                chemistry_number="C1",
                mutation_type="Insertion",
                mutation_length=1,
                mutation_start=22,
                rna_sequence=sequence,
            )
        )
        self.assertEqual(result.aso_length, 18)
        self.assertEqual(result.mutation_start_reversed, len(sequence) - 22)
        self.assertEqual(result.displayed_bases, 37)
        self.assertEqual(result.required_asos, 20)
        self.assertEqual(result.complete_required_asos, 20)
        self.assertEqual(result.coverage_warning, "")
        expected = "".join(
            complement_base(base)
            for base in result.clean_reversed_rna[result.crop_start : result.crop_start + result.aso_length]
        )
        self.assertEqual(result.rows[0].clean_sequence, expected)
        self.assertTrue(any(span.kind == "mutation" for row in result.rows for span in row.display_spans))
        self.assertEqual(
            mutation_header_indexes(result),
            {result.mutation_start_reversed - result.crop_start},
        )
        self.assertEqual(
            header_display_positions_5to3(result)[result.mutation_start_reversed - result.crop_start],
            22,
        )

    def test_variant_microwalk_can_skip_by_step_size(self):
        sequence = "AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA"
        result = generate_design(
            AsoInputs(
                target_gene="GENE",
                snp_identifier="Example",
                chemistry_number="C1",
                mutation_type="Insertion",
                mutation_length=1,
                mutation_start=22,
                microwalk_step_size=2,
                rna_sequence=sequence,
            )
        )
        self.assertEqual(result.required_asos, 10)
        self.assertEqual([row.starting_position for row in result.rows[:4]], [0, 2, 4, 6])
        self.assertEqual([row.row_number for row in result.rows[:4]], [1, 2, 3, 4])
        self.assertTrue(result.rows[1].aso_id.endswith("_ASO_2"))

    def test_rna_gap_markers_are_ignored(self):
        sequence = "AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA"
        marked = "AUGC-UACGUAUGCUACGUAUG_GCAU CGUAUGCUACGUAUGCUACGUA"
        base_inputs = AsoInputs(
            mutation_type="Insertion",
            mutation_length=1,
            mutation_start=22,
            rna_sequence=sequence,
        )
        marked_inputs = AsoInputs(
            mutation_type="Insertion",
            mutation_length=1,
            mutation_start=22,
            rna_sequence=marked,
        )

        plain_result = generate_design(base_inputs)
        marked_result = generate_design(marked_inputs)

        self.assertEqual(marked_result.clean_reversed_rna, plain_result.clean_reversed_rna)
        self.assertEqual(
            [row.clean_sequence for row in marked_result.rows],
            [row.clean_sequence for row in plain_result.rows],
        )

    def test_short_flanking_context_flags_partial_walk(self):
        result = generate_design(
            AsoInputs(
                aso_chemistry="KT777/valeriasen",
                mutation_type="Deletion",
                mutation_length=1,
                mutation_start=17,
                rna_sequence=("A" * 16) + "_" + ("C" * 16),
            )
        )

        self.assertEqual(result.aso_length, 20)
        self.assertEqual(result.required_asos, 13)
        self.assertEqual(result.complete_required_asos, 21)
        self.assertEqual(len(result.rows), 13)
        self.assertIn("Partial walk", result.coverage_warning)
        self.assertIn("Generated 13 of 21", result.coverage_warning)

    def test_substitution_header_highlight_spans_variant_bases(self):
        sequence = "AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA"
        result = generate_design(
            AsoInputs(
                mutation_type="Substitution",
                mutation_length=3,
                mutation_start=21,
                rna_sequence=sequence,
            )
        )
        first_index = result.mutation_start_reversed - result.crop_start
        self.assertEqual(mutation_header_indexes(result), set(range(first_index, first_index + 3)))

    def test_deletion_display_inserts_minimum_visible_gap(self):
        sequence = "AUGCUACGUAUGCUACGUAUGCAUCGUAUGCUACGUAUGCUACGUA"
        result = generate_design(
            AsoInputs(
                mutation_type="Deletion",
                mutation_length=2,
                mutation_start=22,
                rna_sequence=sequence,
            )
        )
        gap_rows = [row for row in result.rows if row.display_spans]
        self.assertTrue(gap_rows)
        first_gap = gap_rows[0].display_spans[0]
        self.assertEqual(first_gap.kind, "gap")
        self.assertEqual(first_gap.end - first_gap.start, 4)
        self.assertIn("    ", gap_rows[0].display_sequence)

    def test_deletion_display_uses_actual_gap_length_when_over_four(self):
        sequence = "AUGCUACGUAUGCUACGUAUGCAUCGUAUGCUACGUAUGCUACGUA"
        result = generate_design(
            AsoInputs(
                mutation_type="Deletion",
                mutation_length=7,
                mutation_start=22,
                rna_sequence=sequence,
            )
        )
        gap = next(row.display_spans[0] for row in result.rows if row.display_spans)
        self.assertEqual(gap.end - gap.start, 7)

    def test_mutation_start_rejects_zero_with_one_based_message(self):
        with self.assertRaisesRegex(ValueError, "first base = 1"):
            generate_design(
                AsoInputs(
                    mutation_type="Insertion",
                    mutation_length=1,
                    mutation_start=0,
                    rna_sequence="AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA",
                )
            )


if __name__ == "__main__":
    unittest.main()
