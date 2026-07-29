import tempfile
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

from aso_logic import AsoInputs, generate_design
from excel_export import export_result_to_xlsx


class ExcelExportTests(unittest.TestCase):
    def test_highlighted_sequence_rich_text_uses_consistent_font_size(self):
        result = generate_design(
            AsoInputs(
                mutation_type="Insertion",
                mutation_length=1,
                mutation_start=22,
                rna_sequence="AUGCUACGUAUGCUACGUAUGGCAUCGUAUGCUACGUAUGCUACGUA",
            )
        )

        with tempfile.TemporaryDirectory() as tmp:
            output = export_result_to_xlsx(result, Path(tmp) / "aso_output.xlsx")
            with zipfile.ZipFile(output) as workbook:
                sheet_xml = workbook.read("xl/worksheets/sheet1.xml")

        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        root = ElementTree.fromstring(sheet_xml)
        highlighted_cell = None
        for cell in root.findall(".//main:c", ns):
            if cell.get("r", "").startswith("C") and cell.findall("main:is/main:r", ns):
                highlighted_cell = cell
                break

        self.assertIsNotNone(highlighted_cell)
        runs = highlighted_cell.findall("main:is/main:r", ns)
        self.assertGreater(len(runs), 1)
        for rich_run in runs:
            run_properties = rich_run.find("main:rPr", ns)
            self.assertIsNotNone(run_properties)
            font = run_properties.find("main:rFont", ns)
            size = run_properties.find("main:sz", ns)
            self.assertIsNotNone(font)
            self.assertIsNotNone(size)
            self.assertEqual(font.get("val"), "Arial")
            self.assertEqual(size.get("val"), "12")

    def test_partial_walk_warning_is_exported(self):
        result = generate_design(
            AsoInputs(
                aso_chemistry="KT777/valeriasen",
                mutation_type="Deletion",
                mutation_length=1,
                mutation_start=17,
                rna_sequence=("A" * 16) + "_" + ("C" * 16),
            )
        )

        with tempfile.TemporaryDirectory() as tmp:
            output = export_result_to_xlsx(result, Path(tmp) / "partial_output.xlsx")
            wb = load_workbook(output)
            ws = wb["ASO Output"]

        self.assertIn("Partial walk", ws["A1"].value)
        self.assertIn("Generated 13 of 19", ws["A1"].value)

    def test_ambiguous_insertion_warning_is_exported(self):
        result = generate_design(
            AsoInputs(
                mutation_type="Insertion",
                mutation_length=1,
                mutation_start=27,
                rna_sequence=("C" * 25) + "AA" + ("G" * 25),
            )
        )

        with tempfile.TemporaryDirectory() as tmp:
            output = export_result_to_xlsx(result, Path(tmp) / "ambiguous_output.xlsx")
            wb = load_workbook(output)
            ws = wb["ASO Output"]

        self.assertIn("Ambiguous insertion placement", ws["A1"].value)
        self.assertIn("26-27", ws["A1"].value)


if __name__ == "__main__":
    unittest.main()
